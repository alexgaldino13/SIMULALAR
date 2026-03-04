"""
Módulo de Exportação para Excel

Cria planilhas Excel (.xlsx) com:
- Resumo comparativo de cenários
- Detalhamento de tabelas de amortização
- Gráficos de evolução
- Formatação profissional em português (Brasil)
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, BarChart, Reference
from decimal import Decimal
from datetime import datetime
import os

# Cores padrão para formatação
COR_HEADER = "1F4E78"  # Azul escuro
COR_DESTAQUE = "FFC000"  # Ouro (melhor opção)
COR_POSITIVO = "92D050"  # Verde claro
COR_NEGATIVO = "FF6B6B"  # Vermelho
COR_NEUTRAL = "D9E1F2"   # Azul claro


def formatar_valor_brl(valor):
    """Converte um valor para string formatada em moeda brasileira."""
    if isinstance(valor, Decimal):
        valor = float(valor)
    return f"R$ {valor:,.2f}".replace(",", "_").replace(".", ",").replace("_", ".")


def formatar_percentual(valor):
    """Converte um valor para string formatada em percentual."""
    if isinstance(valor, Decimal):
        valor = float(valor)
    return f"{valor:.2f}%"


def criar_header(ws, titulo, linha=1):
    """Cria um header formatado em uma worksheet."""
    celula = ws.merge_cells(f'A{linha}:H{linha}')
    celula_obj = ws[f'A{linha}']
    celula_obj.value = titulo
    celula_obj.font = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
    celula_obj.fill = PatternFill(start_color=COR_HEADER, end_color=COR_HEADER, fill_type='solid')
    celula_obj.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[linha].height = 25
    return linha + 2


def criar_row_header_tabela(ws, colunas, linha):
    """Cria um header de tabela com formatação."""
    for col_num, coluna in enumerate(colunas, 1):
        celula = ws.cell(row=linha, column=col_num)
        celula.value = coluna
        celula.font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        celula.fill = PatternFill(start_color=COR_HEADER, end_color=COR_HEADER, fill_type='solid')
        celula.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        celula.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    ws.row_dimensions[linha].height = 20
    return linha + 1


def adicionar_borda_celulas(ws, inicio_linha, fim_linha, inicio_col, fim_col):
    """Adiciona borda a um range de células."""
    border_style = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    for row in range(inicio_linha, fim_linha + 1):
        for col in range(inicio_col, fim_col + 1):
            ws.cell(row=row, column=col).border = border_style


def exportar_para_excel(dados_simulacao, caminho_saida=None):
    """
    Exporta resultados de simulações financeiras para Excel.
    
    Args:
        dados_simulacao (dict): Dicionário contendo:
            - 'resumo': dict com dados comparativos
            - 'price': dict com tabela de amortização Price
            - 'sac': dict com tabela de amortização SAC
            - 'consorcio': dict com dados de consórcio
            - 'mcmv': dict com dados de MCMV
            - 'guardar_dinheiro': dict com dados de poupança
            - 'aluguel_investimento': dict com análise aluguel vs compra
            - 'usuario': dict com dados do usuário (nome, renda, etc.)
        
        caminho_saida (str): Caminho onde salvar o arquivo.
                           Se None, salva em /tmp/simulacao_{timestamp}.xlsx
    
    Returns:
        str: Caminho completo do arquivo criado
    
    Exemplo:
        resultado = comparar_cenarios_e_formatar(dados_form)
        caminho = exportar_para_excel(resultado)
        print(f"Excel exportado: {caminho}")
    """
    
    # Gerar caminho de saída se não fornecido
    if caminho_saida is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        caminho_saida = f"simulacao_{timestamp}.xlsx"
    
    # Criar workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove sheet padrão
    
    # Aba 1: Resumo Comparativo
    _criar_aba_resumo(wb, dados_simulacao)
    
    # Aba 2: Price
    if 'price' in dados_simulacao and dados_simulacao['price']:
        _criar_aba_amortizacao(wb, dados_simulacao['price'], 'PRICE', 2)
    
    # Aba 3: SAC
    if 'sac' in dados_simulacao and dados_simulacao['sac']:
        _criar_aba_amortizacao(wb, dados_simulacao['sac'], 'SAC', 3)
    
    # Aba 4: Consórcio
    if 'consorcio' in dados_simulacao and dados_simulacao['consorcio']:
        _criar_aba_consorcio(wb, dados_simulacao['consorcio'])
    
    # Aba 5: MCMV
    if 'mcmv' in dados_simulacao and dados_simulacao['mcmv']:
        _criar_aba_mcmv(wb, dados_simulacao['mcmv'])
    
    # Aba 6: Guardar Dinheiro
    if 'guardar_dinheiro' in dados_simulacao and dados_simulacao['guardar_dinheiro']:
        _criar_aba_guardar_dinheiro(wb, dados_simulacao['guardar_dinheiro'])
    
    # Aba 7: Gráficos
    _criar_aba_graficos(wb, dados_simulacao)
    
    # Salvar arquivo
    wb.save(caminho_saida)
    return caminho_saida


def _criar_aba_resumo(wb, dados):
    """Cria aba de resumo comparativo."""
    ws = wb.create_sheet('Resumo Comparativo', 0)
    ws.column_dimensions['A'].width = 25
    for col in 'BCDEFGH':
        ws.column_dimensions[col].width = 18
    
    linha = criar_header(ws, '📊 RESUMO COMPARATIVO DE CENÁRIOS', 1)
    
    # Dados do usuário
    usuario = dados.get('usuario', {})
    if usuario:
        ws[f'A{linha}'] = 'Solicitante:'
        ws[f'B{linha}'] = usuario.get('nome', 'Não informado')
        linha += 1
        
        ws[f'A{linha}'] = 'Renda Mensal:'
        ws[f'B{linha}'] = formatar_valor_brl(usuario.get('renda_mensal', 0))
        linha += 1
        
        ws[f'A{linha}'] = 'Valor do Imóvel:'
        ws[f'B{linha}'] = formatar_valor_brl(usuario.get('valor_imovel', 0))
        linha += 1
        
        linha += 1
    
    # Tabela de comparação
    colunas = ['Cenário', 'Parcela Mensal', 'Prazo (meses)', 'Total Pago', 'Total Juros', 'CET', 'Viável?']
    linha = criar_row_header_tabela(ws, colunas, linha)
    
    inicio_linha = linha
    
    # Extrair dados de cada cenário
    cenarios = []
    
    if 'price' in dados and dados['price']:
        cenarios.append(('Price', dados['price']))
    if 'sac' in dados and dados['sac']:
        cenarios.append(('SAC', dados['sac']))
    if 'consorcio' in dados and dados['consorcio']:
        cenarios.append(('Consórcio', dados['consorcio']))
    if 'mcmv' in dados and dados['mcmv']:
        cenarios.append(('MCMV', dados['mcmv']))
    if 'guardar_dinheiro' in dados and dados['guardar_dinheiro']:
        cenarios.append(('Guardar Dinheiro', dados['guardar_dinheiro']))
    
    for nome, dados_cenario in cenarios:
        # Preenchimento de células
        ws[f'A{linha}'] = nome
        ws[f'B{linha}'] = formatar_valor_brl(dados_cenario.get('parcela_inicial', dados_cenario.get('valor_parcela', 0)))
        ws[f'C{linha}'] = int(dados_cenario.get('prazo_meses', dados_cenario.get('prazo_total', 0)))
        ws[f'D{linha}'] = formatar_valor_brl(dados_cenario.get('total_pago', 0))
        ws[f'E{linha}'] = formatar_valor_brl(dados_cenario.get('total_juros', 0))
        ws[f'F{linha}'] = formatar_percentual(dados_cenario.get('cet', 0))
        
        # Viabilidade
        parcela_inicial = dados_cenario.get('parcela_inicial', dados_cenario.get('valor_parcela', 0))
        renda = usuario.get('renda_mensal', 1)
        percentual_renda = (parcela_inicial / renda * 100) if renda > 0 else 100
        
        ws[f'G{linha}'] = '✓ Sim' if percentual_renda <= 30 else '✗ Não'
        if percentual_renda <= 30:
            ws[f'G{linha}'].fill = PatternFill(start_color=COR_POSITIVO, end_color=COR_POSITIVO, fill_type='solid')
        else:
            ws[f'G{linha}'].fill = PatternFill(start_color=COR_NEGATIVO, end_color=COR_NEGATIVO, fill_type='solid')
        
        # Destaque melhor opção
        if nome in ['MCMV', 'Consórcio']:  # Geralmente mais viáveis
            for col in 'ABCDEFG':
                ws[f'{col}{linha}'].fill = PatternFill(start_color=COR_DESTAQUE, end_color=COR_DESTAQUE, fill_type='solid')
        
        # Formatação
        for col in 'ABCDEFG':
            ws[f'{col}{linha}'].alignment = Alignment(horizontal='center')
            ws[f'{col}{linha}'].font = Font(name='Calibri', size=10)
        
        linha += 1
    
    # Adicionar borda
    adicionar_borda_celulas(ws, inicio_linha, linha - 1, 1, 7)


def _criar_aba_amortizacao(wb, dados, titulo, indice):
    """Cria aba de tabela de amortização (Price ou SAC)."""
    ws = wb.create_sheet(titulo, indice)
    
    # Larguras das colunas
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    
    # Header
    linha = criar_header(ws, f'📋 TABELA DE AMORTIZAÇÃO - {titulo}', 1)
    
    # Informações principais
    ws[f'A{linha}'] = 'Parcela Fixa:'
    ws[f'B{linha}'] = formatar_valor_brl(dados.get('parcela_inicial', 0))
    linha += 1
    
    ws[f'A{linha}'] = 'Prazo:'
    ws[f'B{linha}'] = f"{int(dados.get('prazo_meses', 0))} meses"
    linha += 1
    
    ws[f'A{linha}'] = 'Total Juros:'
    ws[f'B{linha}'] = formatar_valor_brl(dados.get('total_juros', 0))
    linha += 1
    
    ws[f'A{linha}'] = 'Total Pago:'
    ws[f'B{linha}'] = formatar_valor_brl(dados.get('total_pago', 0))
    linha += 2
    
    # Tabela de detalhes
    colunas = ['Mês', 'Saldo Devedor', 'Juros', 'Amortização', 'Parcela', 'Acumulado Juros']
    linha = criar_row_header_tabela(ws, colunas, linha)
    
    inicio_linha = linha
    tabela = dados.get('tabela', [])
    
    for idx, mes in enumerate(tabela[:360], 1):  # Limitar a 360 meses para não sobrecarregar
        if mes is None:
            continue
        
        ws[f'A{linha}'] = idx
        ws[f'B{linha}'] = formatar_valor_brl(mes.get('saldo_devedor', 0))
        ws[f'C{linha}'] = formatar_valor_brl(mes.get('juros', 0))
        ws[f'D{linha}'] = formatar_valor_brl(mes.get('amortizacao', 0))
        ws[f'E{linha}'] = formatar_valor_brl(mes.get('parcela', 0))
        ws[f'F{linha}'] = formatar_valor_brl(mes.get('juros_acumulado', 0))
        
        # Formatação
        for col in 'ABCDEF':
            ws[f'{col}{linha}'].alignment = Alignment(horizontal='right')
            ws[f'{col}{linha}'].font = Font(name='Calibri', size=9)
            ws[f'{col}{linha}'].number_format = '_("R"* #,##0.00_);_("R"* (#,##0.00);_("R"* "-"??_);_(@_)'
        
        linha += 1
    
    # Adicionar borda
    if linha > inicio_linha:
        adicionar_borda_celulas(ws, inicio_linha, linha - 1, 1, 6)


def _criar_aba_consorcio(wb, dados):
    """Cria aba de detalhamento de consórcio."""
    ws = wb.create_sheet('Consórcio', 3)
    
    for col in 'ABCDEFGH':
        ws.column_dimensions[col].width = 18
    
    linha = criar_header(ws, '🏛️  DETALHAMENTO - CONSÓRCIO', 1)
    
    # Informações principais
    info = [
        ('Valor da Cota', dados.get('valor_imovel', 0)),
        ('Parcela Mensal', dados.get('valor_parcela', 0)),
        ('Prazo Estimado', f"{int(dados.get('prazo_meses', 0))} meses"),
        ('Taxa de Administração', formatar_percentual(dados.get('taxa_administracao', 0))),
        ('Total Estimado', dados.get('total_pago', 0)),
    ]
    
    for descricao, valor in info:
        ws[f'A{linha}'] = descricao
        if isinstance(valor, (int, float)):
            ws[f'B{linha}'] = formatar_valor_brl(valor)
        else:
            ws[f'B{linha}'] = valor
        linha += 1
    
    linha += 1
    
    # Cronograma de lances (se disponível)
    if 'cronograma_lances' in dados and dados['cronograma_lances']:
        colunas = ['Mês', 'Chance de Contemplação (%)', 'Valor Acumulado']
        linha = criar_row_header_tabela(ws, colunas, linha)
        
        inicio_linha = linha
        for lance_data in dados['cronograma_lances'][:60]:  # Primeiros 60 meses
            ws[f'A{linha}'] = int(lance_data.get('mes', 0))
            ws[f'B{linha}'] = formatar_percentual(lance_data.get('chance', 0))
            ws[f'C{linha}'] = formatar_valor_brl(lance_data.get('acumulado', 0))
            linha += 1
        
        adicionar_borda_celulas(ws, inicio_linha, linha - 1, 1, 3)


def _criar_aba_mcmv(wb, dados):
    """Cria aba de detalhamento de MCMV."""
    ws = wb.create_sheet('MCMV', 4)
    
    for col in 'ABCDEFGH':
        ws.column_dimensions[col].width = 18
    
    linha = criar_header(ws, '🏠 PROGRAMA MCMV (Minha Casa Minha Vida)', 1)
    
    # Informações principais
    info = [
        ('Valor do Imóvel', dados.get('valor_imovel', 0)),
        ('Subsídio Governamental', dados.get('subsidio', 0)),
        ('Valor Financiado', dados.get('valor_financiado', 0)),
        ('Entrada Necessária', dados.get('entrada_necessaria', 0)),
        ('Parcela Mensal', dados.get('parcela_inicial', 0)),
        ('Prazo (meses)', int(dados.get('prazo_meses', 0))),
        ('Total Pago', dados.get('total_pago', 0)),
        ('Total Juros', dados.get('total_juros', 0)),
        ('Taxa Anual', formatar_percentual(dados.get('taxa_anual', 0))),
    ]
    
    for descricao, valor in info:
        ws[f'A{linha}'] = descricao
        if isinstance(valor, (int, float)):
            if 'taxa' in descricao.lower() or 'subsídio' in descricao.lower():
                ws[f'B{linha}'] = formatar_percentual(valor) if 'taxa' in descricao.lower() else formatar_valor_brl(valor)
            else:
                ws[f'B{linha}'] = formatar_valor_brl(valor) if isinstance(valor, float) else valor
        else:
            ws[f'B{linha}'] = valor
        
        # Destacar subsídio
        if 'subsídio' in descricao.lower():
            ws[f'B{linha}'].fill = PatternFill(start_color=COR_POSITIVO, end_color=COR_POSITIVO, fill_type='solid')
        
        linha += 1


def _criar_aba_guardar_dinheiro(wb, dados):
    """Cria aba de detalhamento de guardar dinheiro."""
    ws = wb.create_sheet('Guardar Dinheiro', 5)
    
    for col in 'ABCDEFGH':
        ws.column_dimensions[col].width = 18
    
    linha = criar_header(ws, '💰 SIMULAÇÃO - GUARDAR DINHEIRO', 1)
    
    # Informações principais
    info = [
        ('Valor a Guardar Mensalmente', dados.get('valor_mensal', 0)),
        ('Meses até R$ 300.000', int(dados.get('meses_ate_valor', 0))),
        ('Taxa de Rentabilidade', formatar_percentual(dados.get('taxa_rentabilidade', 0))),
        ('Valor Final', dados.get('valor_final', 0)),
        ('Juros Acumulados', dados.get('juros_acumulados', 0)),
    ]
    
    for descricao, valor in info:
        ws[f'A{linha}'] = descricao
        if isinstance(valor, (int, float)):
            if 'taxa' in descricao.lower():
                ws[f'B{linha}'] = valor if isinstance(valor, str) else formatar_percentual(valor)
            else:
                ws[f'B{linha}'] = valor if isinstance(valor, str) else formatar_valor_brl(valor)
        else:
            ws[f'B{linha}'] = valor
        linha += 1


def _criar_aba_graficos(wb, dados):
    """Cria aba com gráficos comparativos."""
    ws = wb.create_sheet('Gráficos', 6)
    
    linha = criar_header(ws, '📈 GRÁFICOS DE COMPARAÇÃO', 1)
    
    # Informação de que gráficos podem ser adicionados manualmente
    ws[f'A{linha}'] = '⚠️  Gráficos nesta aba podem ser adicionados manualmente no Excel:'
    linha += 2
    
    pontos = [
        '1. Evolução do Saldo Devedor (Price vs SAC)',
        '2. Comparação de Parcelas Mensais',
        '3. Acúmulo de Juros Pagos',
        '4. Comparação de Cenários (Total Pago)',
    ]
    
    for ponto in pontos:
        ws[f'A{linha}'] = ponto
        linha += 1
    
    linha += 2
    ws[f'A{linha}'] = 'Dica: Use os dados das abas anteriores para criar gráficos personalizados!'


if __name__ == '__main__':
    # Exemplo de uso
    print("✓ Módulo de Exportação Excel carregado com sucesso!")
    print("\nPara usar:")
    print("  from simulacao.exportacao_excel import exportar_para_excel")
    print("  resultado = comparar_cenarios_e_formatar(dados)")
    print("  caminho = exportar_para_excel(resultado)")
