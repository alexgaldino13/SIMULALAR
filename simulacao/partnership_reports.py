# simulacao/partnership_reports.py
"""
Sistema de Relatórios de Performance para Parcerias
Gera relatórios gerenciais, ROI e exportações
"""

from django.db.models import Count, Sum, Avg, Q, F
from django.utils import timezone
from datetime import timedelta
import csv
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import BarChart, Reference


class PartnershipReportGenerator:
    """
    Gerador de relatórios para o sistema de parcerias.
    """
    
    def __init__(self, parceiro=None, data_inicio=None, data_fim=None):
        """
        Inicializa o gerador de relatórios.
        
        Args:
            parceiro: Partnership object (opcional, None = todos os parceiros)
            data_inicio: Data de início do período
            data_fim: Data de fim do período
        """
        from .partnership_models import Partnership, Lead
        
        self.parceiro = parceiro
        self.data_inicio = data_inicio or (timezone.now() - timedelta(days=30))
        self.data_fim = data_fim or timezone.now()
        
        # Query base de leads
        self.leads_query = Lead.objects.filter(
            criado_em__gte=self.data_inicio,
            criado_em__lte=self.data_fim
        )
        
        if parceiro:
            self.leads_query = self.leads_query.filter(parceiro=parceiro)
    
    def get_metricas_gerais(self):
        """
        Retorna métricas gerais do período.
        """
        leads = self.leads_query
        
        total_leads = leads.count()
        leads_convertidos = leads.filter(status='CONVERTIDO').count()
        leads_perdidos = leads.filter(status='PERDIDO').count()
        leads_em_andamento = leads.exclude(status__in=['CONVERTIDO', 'PERDIDO', 'INVALIDO']).count()
        
        taxa_conversao = (leads_convertidos / total_leads * 100) if total_leads > 0 else 0
        taxa_perda = (leads_perdidos / total_leads * 100) if total_leads > 0 else 0
        
        # Valores financeiros
        valor_total_negocios = leads.filter(
            status='CONVERTIDO'
        ).aggregate(Sum('valor_negocio'))['valor_negocio__sum'] or 0
        
        comissao_total = leads.filter(
            status='CONVERTIDO'
        ).aggregate(Sum('comissao_gerada'))['comissao_gerada__sum'] or 0
        
        ticket_medio = valor_total_negocios / leads_convertidos if leads_convertidos > 0 else 0
        
        # Tempo médio de conversão
        leads_convertidos_obj = leads.filter(
            status='CONVERTIDO',
            convertido_em__isnull=False
        )
        
        tempo_medio_conversao = 0
        if leads_convertidos_obj.exists():
            tempos = []
            for lead in leads_convertidos_obj:
                if lead.convertido_em and lead.criado_em:
                    dias = (lead.convertido_em - lead.criado_em).days
                    tempos.append(dias)
            tempo_medio_conversao = sum(tempos) / len(tempos) if tempos else 0
        
        return {
            'total_leads': total_leads,
            'leads_convertidos': leads_convertidos,
            'leads_perdidos': leads_perdidos,
            'leads_em_andamento': leads_em_andamento,
            'taxa_conversao': round(taxa_conversao, 2),
            'taxa_perda': round(taxa_perda, 2),
            'valor_total_negocios': float(valor_total_negocios),
            'comissao_total': float(comissao_total),
            'ticket_medio': float(ticket_medio),
            'tempo_medio_conversao_dias': round(tempo_medio_conversao, 1),
        }
    
    def get_leads_por_status(self):
        """
        Retorna distribuição de leads por status.
        """
        return self.leads_query.values('status').annotate(
            total=Count('id')
        ).order_by('-total')
    
    def get_leads_por_origem(self):
        """
        Retorna distribuição de leads por origem.
        """
        return self.leads_query.values('origem').annotate(
            total=Count('id')
        ).order_by('-total')
    
    def get_leads_por_estado(self):
        """
        Retorna distribuição de leads por estado.
        """
        return self.leads_query.values('estado_interesse').annotate(
            total=Count('id')
        ).order_by('-total')
    
    def get_funil_conversao(self):
        """
        Retorna dados do funil de conversão.
        """
        leads = self.leads_query
        
        funil = {
            'total_leads': leads.count(),
            'enviados': leads.filter(status__in=['ENVIADO', 'EM_CONTATO', 'QUALIFICADO', 'NEGOCIACAO', 'CONVERTIDO']).count(),
            'em_contato': leads.filter(status__in=['EM_CONTATO', 'QUALIFICADO', 'NEGOCIACAO', 'CONVERTIDO']).count(),
            'qualificados': leads.filter(status__in=['QUALIFICADO', 'NEGOCIACAO', 'CONVERTIDO']).count(),
            'em_negociacao': leads.filter(status__in=['NEGOCIACAO', 'CONVERTIDO']).count(),
            'convertidos': leads.filter(status='CONVERTIDO').count(),
        }
        
        # Calcula taxas de conversão entre etapas
        funil['taxa_envio'] = (funil['enviados'] / funil['total_leads'] * 100) if funil['total_leads'] > 0 else 0
        funil['taxa_contato'] = (funil['em_contato'] / funil['enviados'] * 100) if funil['enviados'] > 0 else 0
        funil['taxa_qualificacao'] = (funil['qualificados'] / funil['em_contato'] * 100) if funil['em_contato'] > 0 else 0
        funil['taxa_negociacao'] = (funil['em_negociacao'] / funil['qualificados'] * 100) if funil['qualificados'] > 0 else 0
        funil['taxa_conversao_final'] = (funil['convertidos'] / funil['em_negociacao'] * 100) if funil['em_negociacao'] > 0 else 0
        
        return funil
    
    def get_roi_parceiro(self):
        """
        Calcula o ROI do parceiro.
        """
        if not self.parceiro:
            return None
        
        leads = self.leads_query
        
        # Custo: valor pago por lead
        custo_total = leads.count() * float(self.parceiro.valor_por_lead)
        
        # Receita: comissões geradas
        receita_total = float(
            leads.filter(status='CONVERTIDO').aggregate(
                Sum('comissao_gerada')
            )['comissao_gerada__sum'] or 0
        )
        
        # ROI = (Receita - Custo) / Custo * 100
        roi = ((receita_total - custo_total) / custo_total * 100) if custo_total > 0 else 0
        
        return {
            'custo_total': custo_total,
            'receita_total': receita_total,
            'lucro': receita_total - custo_total,
            'roi_percentual': round(roi, 2),
        }
    
    def get_evolucao_mensal(self, meses=12):
        """
        Retorna evolução mensal dos leads.
        """
        from .partnership_models import Lead
        from django.db.models.functions import TruncMonth
        
        data_inicio = timezone.now() - timedelta(days=meses*30)
        
        query = Lead.objects.filter(
            criado_em__gte=data_inicio
        )
        
        if self.parceiro:
            query = query.filter(parceiro=self.parceiro)
        
        evolucao = query.annotate(
            mes=TruncMonth('criado_em')
        ).values('mes').annotate(
            total=Count('id'),
            convertidos=Count('id', filter=Q(status='CONVERTIDO')),
            perdidos=Count('id', filter=Q(status='PERDIDO')),
        ).order_by('mes')
        
        return list(evolucao)
    
    def get_top_parceiros(self, limit=10):
        """
        Retorna os top parceiros por performance.
        """
        from .partnership_models import Partnership
        
        parceiros = Partnership.objects.filter(
            status='ATIVO'
        ).annotate(
            total_leads=Count('leads', filter=Q(
                leads__criado_em__gte=self.data_inicio,
                leads__criado_em__lte=self.data_fim
            )),
            leads_convertidos=Count('leads', filter=Q(
                leads__criado_em__gte=self.data_inicio,
                leads__criado_em__lte=self.data_fim,
                leads__status='CONVERTIDO'
            )),
            valor_total=Sum('leads__valor_negocio', filter=Q(
                leads__criado_em__gte=self.data_inicio,
                leads__criado_em__lte=self.data_fim,
                leads__status='CONVERTIDO'
            )),
        ).order_by('-valor_total')[:limit]
        
        return parceiros
    
    def exportar_csv(self):
        """
        Exporta relatório em formato CSV.
        """
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Cabeçalho
        writer.writerow([
            'Nome',
            'Email',
            'Telefone',
            'Valor Imóvel',
            'Cidade',
            'Estado',
            'Status',
            'Origem',
            'Data Criação',
            'Data Conversão',
            'Valor Negócio',
            'Comissão',
        ])
        
        # Dados
        for lead in self.leads_query.select_related('parceiro'):
            writer.writerow([
                lead.nome_completo,
                lead.email,
                lead.telefone,
                f"R$ {lead.valor_imovel:,.2f}",
                lead.cidade_interesse or '',
                lead.estado_interesse or '',
                lead.get_status_display(),
                lead.get_origem_display(),
                lead.criado_em.strftime('%d/%m/%Y'),
                lead.convertido_em.strftime('%d/%m/%Y') if lead.convertido_em else '',
                f"R$ {lead.valor_negocio:,.2f}" if lead.valor_negocio else '',
                f"R$ {lead.comissao_gerada:,.2f}" if lead.comissao_gerada else '',
            ])
        
        return output.getvalue()
    
    def exportar_excel(self):
        """
        Exporta relatório em formato Excel com gráficos.
        """
        wb = Workbook()
        
        # Aba 1: Métricas Gerais
        ws1 = wb.active
        ws1.title = "Métricas Gerais"
        
        # Título
        ws1['A1'] = 'RELATÓRIO DE PERFORMANCE - PARCERIAS'
        ws1['A1'].font = Font(size=16, bold=True)
        ws1.merge_cells('A1:D1')
        
        ws1['A2'] = f'Período: {self.data_inicio.strftime("%d/%m/%Y")} a {self.data_fim.strftime("%d/%m/%Y")}'
        ws1.merge_cells('A2:D2')
        
        # Métricas
        metricas = self.get_metricas_gerais()
        row = 4
        
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        
        ws1[f'A{row}'] = 'Métrica'
        ws1[f'B{row}'] = 'Valor'
        ws1[f'A{row}'].fill = header_fill
        ws1[f'B{row}'].fill = header_fill
        ws1[f'A{row}'].font = header_font
        ws1[f'B{row}'].font = header_font
        
        row += 1
        for key, value in metricas.items():
            ws1[f'A{row}'] = key.replace('_', ' ').title()
            ws1[f'B{row}'] = value
            row += 1
        
        # Aba 2: Leads Detalhados
        ws2 = wb.create_sheet("Leads")
        
        # Cabeçalho
        headers = ['Nome', 'Email', 'Telefone', 'Valor Imóvel', 'Status', 'Data Criação']
        for col, header in enumerate(headers, 1):
            cell = ws2.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
        
        # Dados
        for row, lead in enumerate(self.leads_query, 2):
            ws2.cell(row=row, column=1, value=lead.nome_completo)
            ws2.cell(row=row, column=2, value=lead.email)
            ws2.cell(row=row, column=3, value=lead.telefone)
            ws2.cell(row=row, column=4, value=float(lead.valor_imovel))
            ws2.cell(row=row, column=5, value=lead.get_status_display())
            ws2.cell(row=row, column=6, value=lead.criado_em.strftime('%d/%m/%Y'))
        
        # Ajusta largura das colunas
        for ws in [ws1, ws2]:
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Salva em memória
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output.getvalue()
    
    def gerar_relatorio_completo(self):
        """
        Gera um relatório completo com todas as métricas.
        """
        return {
            'periodo': {
                'inicio': self.data_inicio.strftime('%d/%m/%Y'),
                'fim': self.data_fim.strftime('%d/%m/%Y'),
            },
            'parceiro': self.parceiro.nome if self.parceiro else 'Todos',
            'metricas_gerais': self.get_metricas_gerais(),
            'leads_por_status': list(self.get_leads_por_status()),
            'leads_por_origem': list(self.get_leads_por_origem()),
            'leads_por_estado': list(self.get_leads_por_estado()),
            'funil_conversao': self.get_funil_conversao(),
            'roi': self.get_roi_parceiro() if self.parceiro else None,
            'evolucao_mensal': self.get_evolucao_mensal(),
        }
